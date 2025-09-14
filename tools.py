import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from utils import get_embedding_model, get_knoweledge_base
from langchain_core.tools import tool


file_path = "leads.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "capture_timestamp", "full_name", "work_email", "company", 
        "qualification_score", "lead_quality", 
        "pages_visited", "questions_asked", "time_to_capture",
        "scroll_percentage", "session_id"
    ])
    wb.save(file_path)

@tool
def get_customer_support() -> str:
    """
    Retrieve customer support contact details.

    Returns:
        str: The customer support contact information, including email and 
        support hours, so that users can easily reach out for assistance.
    """
    support_info = (
        "📞 FinKraft Customer Support\n"
        "Email: support@flipkraft.com\n"
        "Phone: +91-8000-123-456\n"
        "Support Hours: Mon–Sat, 10 AM – 8 PM IST\n"
        "Knowledge Base: https://finkraft.com/help\n"
        "Live Chat: Available on the website"
    )
    return support_info


@tool
def get_context(query: str):
    """
    Retrieve relevant SaaS product and FAQ information for a user's query.

    This tool searches the SaaS knowledge base, including features, pricing, 
    security, compliance, API limits, and common customer questions, to provide 
    context that helps the chatbot answer user queries accurately and in a 
    business-focused manner.

    Args:
        query (str): A user's question or request related to the SaaS platform, 
            such as SOC 2 compliance, data residency, integrations, pricing tiers, 
            or feature limits.

    Returns:
        str: A string containing the most relevant information from the SaaS 
        knowledge base to assist the chatbot in generating precise and helpful 
        responses.
    """
    embedder = get_embedding_model()
    index, answers = get_knoweledge_base()
    query_vec = embedder.encode([query], convert_to_numpy=True)
    D, I = index.search(query_vec, 2)
    return answers[I[0][0]]


@tool
def save_lead_to_excel(user_input: str) -> str:
    """
    Extracts lead/contact details (name, email, company) from user input
    and saves them into an Excel file (leads.xlsx) in the proper format.
    """
    # --- Extract with regex ---
    name_match = re.search(r"(?:name|my name is)\s*[:\-]?\s*([A-Za-z ]+)", user_input, re.IGNORECASE)
    email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}", user_input)
    company_match = re.search(r"(?:company|org|organization)\s*[:\-]?\s*([A-Za-z0-9 &]+)", user_input, re.IGNORECASE)

    lead = {
        "capture_timestamp": datetime.now().isoformat(),
        "full_name": name_match.group(1).strip() if name_match else "",
        "work_email": email_match.group(0).strip() if email_match else "",
        "company": company_match.group(1).strip() if company_match else "",
        "qualification_score": 100,   
        "lead_quality": "Very_hot",      
        "pages_visited": "/home",    
        "questions_asked": 0,
        "time_to_capture": 0,
        "scroll_percentage": 0,      
        "session_id": ""             
    }


    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([
        lead["capture_timestamp"], lead["full_name"], lead["work_email"], lead["company"],
        lead["qualification_score"], lead["lead_quality"],
        lead["pages_visited"], lead["questions_asked"], lead["time_to_capture"],
        lead["scroll_percentage"], lead["session_id"]
    ])
    wb.save(file_path)

    return f"Lead saved: {lead['full_name']} | {lead['work_email']} | {lead['company']}"